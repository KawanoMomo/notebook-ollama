from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

from core.exceptions import AppError, ErrorCode
from core.ingestion.parsers import register
from core.ingestion.types import ParsedDocument, ParsedSection


def _sheet_to_csv(rows) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if v is None else str(v) for v in row])
    return buf.getvalue().rstrip("\n")


class XlsxParser:
    kind = "xlsx"

    def parse_bytes(self, data: bytes, *, source_hint: str | None = None) -> ParsedDocument:
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        except Exception as exc:
            raise AppError(
                ErrorCode.INGESTION_PARSE_FAILED, "xlsx parse failed", detail=str(exc)
            ) from exc

        sections: list[ParsedSection] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            text = _sheet_to_csv(rows)
            if not text.strip():
                continue
            sections.append(
                ParsedSection(
                    text=text,
                    page=None,
                    heading_path=[sheet_name],
                    ord=idx,
                )
            )
        wb.close()
        if not sections:
            raise AppError(ErrorCode.INGESTION_PARSE_FAILED, "no content in xlsx")
        return ParsedDocument(title=source_hint or "workbook.xlsx", sections=sections)


register(XlsxParser())
